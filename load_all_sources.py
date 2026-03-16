import os, sys, logging
import psycopg2
from psycopg2.extras import execute_batch
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

DB_CONFIG = {
    "host":     os.getenv("PG_HOST",     "localhost"),
    "port":     int(os.getenv("PG_PORT", "5433")),
    "dbname":   os.getenv("PG_DB",       "dw_edm"),
    "user":     os.getenv("PG_USER",     "edm_user"),
    "password": os.getenv("PG_PASSWORD", "edm_pass"),
}

BATCH_SIZE = 500

def get_conn():
    return psycopg2.connect(**DB_CONFIG)

def run():
    files = ["SOURCE 01 CoreBanking Debt.xlsx", "SOURCE 02 EWallet.xlsx", "SOURCE 03 Payment.xlsx"]
    conn = get_conn()
    
    try:
        for f in files:
            if not os.path.exists(f):
                log.warning("File not found: %s", f)
                continue
            
            log.info("Processing file: %s", f)
            wb = load_workbook(f, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                if 'README' in sheet_name.upper():
                    continue
                
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) <= 1:
                    continue
                
                headers = [str(h).lower().strip() for h in rows[0]]
                records = []
                for row in rows[1:]:
                    if not any(x is not None for x in row): continue
                    # Cast datetime/date to string if needed, psycopg2 handles usually, but can be safe.
                    records.append(tuple(row))
                    
                table_name = sheet_name # e.g. corebank.customer
                
                if table_name == 'ewallet.balance':
                    headers = ['account_id', 'current_balance', 'available_balance', 'frozen_amount', 'last_updated']
                    records = [row[:5] for row in records]
                else:
                    # filter out totally empty columns if any
                    valid_cols = [i for i, h in enumerate(headers) if h and h.strip() != '']
                    headers = [headers[i] for i in valid_cols]
                    records = [tuple(r[i] for i in valid_cols) for r in records]
                
                cur = conn.cursor()
                log.info("Truncating table %s ...", table_name)
                cur.execute(f"TRUNCATE {table_name} RESTART IDENTITY CASCADE")
                
                cols_str = ",".join(headers)
                placeholders = ",".join(["%s"] * len(headers))
                sql = f"INSERT INTO {table_name} ({cols_str}) VALUES ({placeholders})"
                
                execute_batch(cur, sql, records, page_size=BATCH_SIZE)
                conn.commit()
                log.info("Finished %s: inserted %d rows", table_name, len(records))
                
            wb.close()
            
    except Exception as e:
        conn.rollback()
        log.error("FAILED during file %s: %s", f, e)
        raise
    finally:
        conn.close()

if __name__ == "__main__":
    run()
