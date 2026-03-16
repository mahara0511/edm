"""
Load E-Wallet & Payment sample data từ Excel → DW_EDM
Chạy SAU KHI đã load file gốc (CyberHive_DebtDW_Dataset.xlsx)
"""
import os, sys, logging
import psycopg2
from psycopg2.extras import execute_batch
from openpyxl import load_workbook
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)
DB_CONFIG = {
    "host":     os.getenv("PG_HOST",     "localhost"),
    "port":     int(os.getenv("PG_PORT", "5433")),
    "dbname":   os.getenv("PG_DB",       "dw_edm"),
    "user":     os.getenv("PG_USER",     "edm_user"),
    "password": os.getenv("PG_PASSWORD", "edm_pass"),
}
BATCH = 500
def conn(): return psycopg2.connect(**DB_CONFIG)
def read_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
def to_float(v):
    try: return float(v) if v is not None else None
    except: return None
def to_int(v):
    try: return int(v) if v is not None else None
    except: return None
def load_dim_merchant(c, rows):
    log.info("Loading dim_merchant: %d rows ...", len(rows))
    cur = c.cursor()
    cur.execute("TRUNCATE dw.dim_merchant RESTART IDENTITY CASCADE")
    execute_batch(cur, """
        INSERT INTO dw.dim_merchant
            (merchant_key, merchant_code, merchant_name, merchant_category, region, branch_code, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
    """, [(to_int(r["merchant_key"]),r["merchant_code"],r["merchant_name"],
           r["merchant_category"],r["region"],r["branch_code"],r["status"]) for r in rows], page_size=BATCH)
    c.commit()
    log.info("dim_merchant: %d rows inserted", len(rows))
def load_dim_channel(c, rows):
    log.info("Loading dim_channel: %d rows ...", len(rows))
    cur = c.cursor()
    cur.execute("TRUNCATE dw.dim_channel RESTART IDENTITY CASCADE")
    execute_batch(cur, """
        INSERT INTO dw.dim_channel
            (channel_key, channel_code, channel_name, channel_type, product_line)
        VALUES (%s,%s,%s,%s,%s)
    """, [(to_int(r["channel_key"]),r["channel_code"],r["channel_name"],
           r["channel_type"],r["product_line"]) for r in rows], page_size=BATCH)
    c.commit()
    log.info("dim_channel: %d rows inserted", len(rows))
def load_fact_ewallet(c, rows):
    log.info("Loading fact_ewallet_transaction: %d rows ...", len(rows))
    cur = c.cursor()
    cur.execute("TRUNCATE dw.fact_ewallet_transaction RESTART IDENTITY CASCADE")
    execute_batch(cur, """
        INSERT INTO dw.fact_ewallet_transaction
            (fact_id, customer_key, branch_key, time_key, channel_key,
             txn_count, txn_amount, fee_amount, net_amount,
             txn_type, txn_status, closing_balance)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, [(to_int(r["fact_id"]), to_int(r["customer_key"]), to_int(r["branch_key"]),
           to_int(r["time_key"]), to_int(r["channel_key"]), to_int(r["txn_count"]),
           to_float(r["txn_amount"]), to_float(r["fee_amount"]), to_float(r["net_amount"]),
           r["txn_type"], r["txn_status"], to_float(r["closing_balance"])) for r in rows], page_size=BATCH)
    c.commit()
    log.info("fact_ewallet_transaction: %d rows inserted", len(rows))
def load_fact_payment(c, rows):
    log.info("Loading fact_payment_transaction: %d rows ...", len(rows))
    cur = c.cursor()
    cur.execute("TRUNCATE dw.fact_payment_transaction RESTART IDENTITY CASCADE")
    execute_batch(cur, """
        INSERT INTO dw.fact_payment_transaction
            (fact_id, customer_key, merchant_key, branch_key, time_key, channel_key,
             txn_count, txn_amount, fee_amount, net_amount, refund_amount,
             payment_method, txn_status, merchant_category)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, [(to_int(r["fact_id"]), to_int(r["customer_key"]), to_int(r["merchant_key"]),
           to_int(r["branch_key"]), to_int(r["time_key"]), to_int(r["channel_key"]),
           to_int(r["txn_count"]), to_float(r["txn_amount"]), to_float(r["fee_amount"]),
           to_float(r["net_amount"]), to_float(r["refund_amount"]),
           r["payment_method"], r["txn_status"], r["merchant_category"]) for r in rows], page_size=BATCH)
    c.commit()
    log.info("fact_payment_transaction: %d rows inserted", len(rows))
def run(path):
    log.info("=== LOAD EWALLET & PAYMENT START ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    merchants = read_sheet(wb, "DIM_MERCHANT")
    channels  = read_sheet(wb, "DIM_CHANNEL")
    ew_facts  = read_sheet(wb, "FACT_EWALLET_TRANSACTION")
    pay_facts = read_sheet(wb, "FACT_PAYMENT_TRANSACTION")
    wb.close()
    c = conn()
    try:
        load_dim_merchant(c, merchants)
        load_dim_channel(c, channels)
        load_fact_ewallet(c, ew_facts)
        load_fact_payment(c, pay_facts)
        log.info("=== LOAD COMPLETE ===")
    except Exception as e:
        c.rollback(); log.error("FAILED: %s", e); raise
    finally:
        c.close()
if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "CyberHive_EWallet_Payment_Dataset.xlsx"
    run(path)