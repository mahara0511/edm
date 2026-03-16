"""
Load sample data từ Excel (CyberHive_DebtDW_Dataset.xlsx) 
thẳng vào DW_EDM schema dw.* (bypass source tables)
"""
import os, sys, logging
import psycopg2
from psycopg2.extras import execute_batch
from openpyxl import load_workbook
from datetime import date, datetime
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)
DB_CONFIG = {
    "host":     os.getenv("PG_HOST",     "localhost"),
    "port":     int(os.getenv("PG_PORT", "5433")),
    "dbname":   os.getenv("PG_DB",       "dw_edm"),
    "user":     os.getenv("PG_USER",     "edm_user"),
    "password": os.getenv("PG_PASSWORD", "edm_pass"),
}
BATCH_SIZE = 500
def get_conn():
    return psycopg2.connect(**DB_CONFIG)
def read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).lower().strip() if h else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]
def to_date(val):
    if val is None: return None
    if isinstance(val, (date, datetime)): return val if isinstance(val, date) else val.date()
    try: return date.fromisoformat(str(val))
    except: return None
def to_float(val):
    try: return float(val) if val is not None else None
    except: return None
def to_int(val):
    try: return int(val) if val is not None else None
    except: return None
def load_dim_customer(conn, rows):
    log.info("Loading dim_customer: %d rows ...", len(rows))
    cur = conn.cursor()
    cur.execute("TRUNCATE dw.dim_customer RESTART IDENTITY CASCADE")
    records = [(
        r["customer_key"], r["cif_number"], r["full_name"],
        r["national_id"], r["phone_number"], r["current_address"], r["email"],
        date.today(), None, True
    ) for r in rows]
    execute_batch(cur, """
        INSERT INTO dw.dim_customer
            (customer_key, cif_number, full_name, national_id, phone_number, current_address, email,
             effective_date, expiry_date, is_current)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, records, page_size=BATCH_SIZE)
    conn.commit()
    log.info("dim_customer: %d rows inserted", len(records))
def load_dim_branch(conn, rows):
    log.info("Loading dim_branch: %d rows ...", len(rows))
    cur = conn.cursor()
    cur.execute("TRUNCATE dw.dim_branch RESTART IDENTITY CASCADE")
    records = [(r["branch_key"], r["branch_code"], r["branch_name"], r["region_name"]) for r in rows]
    execute_batch(cur, """
        INSERT INTO dw.dim_branch (branch_key, branch_code, branch_name, region_name)
        VALUES (%s,%s,%s,%s)
    """, records, page_size=BATCH_SIZE)
    conn.commit()
    log.info("dim_branch: %d rows inserted", len(records))
def load_dim_product(conn, rows):
    log.info("Loading dim_product: %d rows ...", len(rows))
    cur = conn.cursor()
    cur.execute("TRUNCATE dw.dim_product RESTART IDENTITY CASCADE")
    records = [(
        r["product_key"], r["contract_number"], r["product_type"],
        to_float(r["interest_rate"]), to_date(r["maturity_date"])
    ) for r in rows]
    execute_batch(cur, """
        INSERT INTO dw.dim_product (product_key, contract_number, product_type, interest_rate, maturity_date)
        VALUES (%s,%s,%s,%s,%s)
    """, records, page_size=BATCH_SIZE)
    conn.commit()
    log.info("dim_product: %d rows inserted", len(records))
def load_dim_time(conn, rows):
    log.info("Loading dim_time: %d rows ...", len(rows))
    cur = conn.cursor()
    cur.execute("TRUNCATE dw.dim_time RESTART IDENTITY CASCADE")
    records = [(
        r["time_key"], to_date(r["full_date"]),
        to_int(r["day"]), to_int(r["month"]),
        to_int(r["quarter"]), to_int(r["year"]),
        r["day_name"], r["month_name"], r["is_weekend"]
    ) for r in rows]
    execute_batch(cur, """
        INSERT INTO dw.dim_time
            (time_key, full_date, day, month, quarter, year, day_name, month_name, is_weekend)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, records, page_size=BATCH_SIZE)
    conn.commit()
    log.info("dim_time: %d rows inserted", len(records))
def load_fact(conn, rows):
    log.info("Loading fact_debt_status: %d rows ...", len(rows))
    cur = conn.cursor()
    cur.execute("TRUNCATE dw.fact_debt_status RESTART IDENTITY CASCADE")
    records = [(
        r["fact_id"],
        to_int(r["customer_key"]), to_int(r["branch_key"]),
        to_int(r["product_key"]),  to_int(r["time_key"]),
        to_float(r["principal_amount"]), to_float(r["interest_amount"]),
        to_float(r["total_outstanding"]), to_int(r["days_past_due"]),
        r["risk_bucket"]
    ) for r in rows]
    execute_batch(cur, """
        INSERT INTO dw.fact_debt_status
            (fact_id, customer_key, branch_key, product_key, time_key,
             principal_amount, interest_amount, total_outstanding, days_past_due, risk_bucket)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, records, page_size=BATCH_SIZE)
    conn.commit()
    log.info("fact_debt_status: %d rows inserted", len(records))
def run(excel_path):
    log.info("=== LOAD FROM EXCEL START ===")
    log.info("Reading: %s", excel_path)
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    customers = read_sheet(wb, "DIM_CUSTOMER")
    branches  = read_sheet(wb, "DIM_BRANCH")
    products  = read_sheet(wb, "DIM_PRODUCT")
    times     = read_sheet(wb, "DIM_TIME")
    facts     = read_sheet(wb, "FACT_DEBT_STATUS")
    wb.close()
    conn = get_conn()
    try:
        load_dim_customer(conn, customers)
        load_dim_branch(conn, branches)
        load_dim_product(conn, products)
        load_dim_time(conn, times)
        load_fact(conn, facts)
        log.info("=== LOAD COMPLETE ===")
    except Exception as e:
        conn.rollback()
        log.error("FAILED: %s", e)
        raise
    finally:
        conn.close()
if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "CyberHive_DebtDW_Dataset.xlsx"
    run(path)